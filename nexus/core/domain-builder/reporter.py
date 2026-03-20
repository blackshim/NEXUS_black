"""
Report Generator

Generates weekly/monthly reports and performs monthly JSON -> Excel export.

Weekly Report:
- Knowledge DB status (total items, by source)
- This week's activity (log count, result classification)
- Promotion candidates
- Top/bottom success rates

Monthly Report:
- Weekly report + JSON -> Excel export

Usage:
    report = generate_weekly_report(config)
    print(report)
"""

import json
import subprocess
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

KST = timezone(timedelta(hours=9))
PYTHON = sys.executable


def generate_weekly_report(
    knowledge_json_path: str,
    log_dir: str,
    domain_name: str = ""
) -> str:
    """Generates a weekly report in markdown."""
    now = datetime.now(KST)

    # Knowledge DB status
    knowledge_stats = _get_knowledge_stats(knowledge_json_path)

    # Log status
    log_stats = _get_log_stats(log_dir)

    # Promotion candidates
    candidates = _get_promotion_candidates(knowledge_json_path)

    # Assemble report
    report = f"# NEXUS Weekly Report\n"
    report += f"> Domain: {domain_name or '?'} | {now.strftime('%Y-%m-%d %H:%M')}\n\n"

    report += "## Knowledge DB Status\n"
    report += f"- Total items: {knowledge_stats['total']}\n"
    for src, cnt in knowledge_stats.get("by_source", {}).items():
        report += f"  - {src}: {cnt}\n"
    report += "\n"

    report += "## This Week's Activity\n"
    report += f"- Conversation logs: {log_stats['total']}\n"
    if log_stats.get("by_category"):
        for cat, cnt in log_stats["by_category"].items():
            report += f"  - {cat}: {cnt}\n"
    report += "\n"

    if candidates:
        report += f"## Promotion Candidates ({len(candidates)})\n"
        for c in candidates:
            s = c.get("usage_stats", {})
            report += f"- **{c.get('id')}**: {c.get('description', '')[:50]}"
            report += f" (used {s.get('suggested', 0)} times, success rate {s.get('success_rate', 0):.0%})\n"
    else:
        report += "## Promotion Candidates\nNone\n"

    report += "\n"

    # Top/bottom success rates
    usage_items = knowledge_stats.get("usage_items", [])
    if usage_items:
        report += "## Top 3 by Success Rate\n"
        top = sorted(usage_items, key=lambda x: -x.get("success_rate", 0))[:3]
        for item in top:
            report += f"- {item['id']}: {item.get('description', '')[:40]} — {item['success_rate']:.0%}\n"

        low = [i for i in usage_items if i.get("success_rate", 0) < 0.5]
        if low:
            report += "\n## Low Success Rate (below 50%)\n"
            for item in low[:3]:
                report += f"- {item['id']}: {item.get('description', '')[:40]} — {item['success_rate']:.0%}\n"

    return report


def generate_monthly_report(
    knowledge_json_path: str,
    log_dir: str,
    exports_dir: str,
    domain_name: str = ""
) -> dict:
    """Monthly report + JSON -> Excel export."""
    # Include weekly report
    report = generate_weekly_report(knowledge_json_path, log_dir, domain_name)
    report = report.replace("Weekly Report", "Monthly Report")

    # Excel export
    now = datetime.now(KST)
    export_path = str(Path(exports_dir) / f"domain_knowledge_{now.strftime('%Y%m')}.xlsx")

    # Same logic as domain-export MCP, but here executed directly via subprocess
    export_result = _export_to_excel(knowledge_json_path, export_path)

    report += f"\n## Excel Export\n"
    if export_result.get("status") == "ok":
        report += f"- File: {export_result['output_path']}\n"
        report += f"- Items: {export_result['total_items']}\n"
    else:
        report += f"- Failed: {export_result.get('error', '?')}\n"

    return {
        "report": report,
        "export_result": export_result
    }


# =======================================
# Internal Helpers
# =======================================

def _get_knowledge_stats(knowledge_json_path: str) -> dict:
    path = Path(knowledge_json_path)
    if not path.exists():
        return {"total": 0, "by_source": {}, "usage_items": []}

    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    items = data.get("items", [])
    by_source = {}
    usage_items = []

    for item in items:
        src = item.get("source", "unknown")
        by_source[src] = by_source.get(src, 0) + 1

        stats = item.get("usage_stats", {})
        if stats.get("suggested", 0) > 0:
            usage_items.append({
                "id": item.get("id"),
                "description": item.get("description", ""),
                "success_rate": stats.get("success_rate", 0),
                "suggested": stats.get("suggested", 0)
            })

    return {"total": len(items), "by_source": by_source, "usage_items": usage_items}


def _get_log_stats(log_dir: str) -> dict:
    path = Path(log_dir)
    if not path.exists():
        return {"total": 0, "by_category": {}}

    total = 0
    by_category = {}

    for log_file in path.glob("*.jsonl"):
        with open(log_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    total += 1
                    cat = entry.get("category", "unknown")
                    by_category[cat] = by_category.get(cat, 0) + 1
                except json.JSONDecodeError:
                    continue

    return {"total": total, "by_category": by_category}


def _get_promotion_candidates(knowledge_json_path: str) -> list[dict]:
    path = Path(knowledge_json_path)
    if not path.exists():
        return []

    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    candidates = []
    for item in data.get("items", []):
        if item.get("source") != "conversation":
            continue
        stats = item.get("usage_stats", {})
        if stats.get("suggested", 0) >= 5 and stats.get("success_rate", 0) >= 0.8:
            candidates.append(item)

    return candidates


def _export_to_excel(knowledge_json_path: str, output_path: str) -> dict:
    """JSON -> Excel export (subprocess)."""
    try:
        # Similar to converter module logic but simplified version
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        script_args = json.dumps({
            "json_path": knowledge_json_path,
            "output_path": output_path
        })

        # Simplified version without reusing domain-export's _EXPORT_SCRIPT
        script = '''
import json, sys, os
args = json.loads(sys.argv[1])
try:
    import openpyxl
    with open(args["json_path"], "r", encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("items", [])
    if not items:
        json.dump({"error": "no items"}, sys.stdout)
        sys.exit(0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "domain_knowledge"
    # Collect all keys
    keys = []
    for item in items:
        for k in item:
            if k not in keys and k != "usage_stats":
                keys.append(k)
    keys.extend(["suggested","resolved","failed","success_rate"])
    for ci, k in enumerate(keys, 1):
        ws.cell(row=1, column=ci, value=k)
    for ri, item in enumerate(items, 2):
        for ci, k in enumerate(keys, 1):
            if k in ("suggested","resolved","failed","success_rate"):
                v = item.get("usage_stats",{}).get(k, 0)
            elif k == "tags":
                v = ", ".join(item.get(k,[])) if isinstance(item.get(k), list) else str(item.get(k,""))
            else:
                v = item.get(k, "")
                if isinstance(v, (list, dict)):
                    v = json.dumps(v, ensure_ascii=False)
            ws.cell(row=ri, column=ci, value=v)
    os.makedirs(os.path.dirname(args["output_path"]), exist_ok=True)
    wb.save(args["output_path"])
    json.dump({"status":"ok","output_path":args["output_path"],"total_items":len(items)}, sys.stdout, ensure_ascii=False)
except Exception as e:
    json.dump({"error":str(e)}, sys.stdout, ensure_ascii=False)
'''
        proc = subprocess.run(
            [PYTHON, "-c", script, script_args],
            capture_output=True, text=True, timeout=60,
            stdin=subprocess.DEVNULL
        )
        if proc.returncode != 0:
            return {"error": proc.stderr.strip()[:200]}
        return json.loads(proc.stdout)

    except Exception as e:
        return {"error": str(e)}
