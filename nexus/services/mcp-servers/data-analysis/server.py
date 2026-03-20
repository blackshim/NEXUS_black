"""
NEXUS data-analysis MCP server -- Excel/CSV data analysis.

Resolves the issue of openpyxl blocking the async event loop
by separating file I/O into a subprocess.
"""

import json
import logging
import os
import subprocess
import sys
from pathlib import Path

try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    class FastMCP:
        def __init__(self, name): self.name = name
        def tool(self, *a, **kw):
            def dec(fn): return fn
            return dec
        def run(self, **kw): pass

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("nexus.mcp.data-analysis")

DOCS_PATH = os.environ.get("DOCS_PATH", "/documents")
PYTHON = sys.executable

mcp = FastMCP("nexus-data-analysis")

# === Helper script (executed via subprocess) ===
_READ_SCRIPT = r'''
import json, sys, os
file_path = sys.argv[1]
sheet_filter = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] != "" else None

if file_path.lower().endswith(".csv"):
    import csv
    rows = []
    for enc in ["utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            with open(file_path, "r", encoding=enc) as f:
                rows = list(csv.reader(f))
            break
        except: pass
    name = os.path.basename(file_path)
    if not rows:
        json.dump({"file": name, "sheets": {}}, sys.stdout, ensure_ascii=False)
    else:
        json.dump({"file": name, "sheets": {"Sheet1": {"header": rows[0], "rows": rows[1:], "row_count": len(rows)-1, "col_count": len(rows[0])}}}, sys.stdout, ensure_ascii=False)
    sys.exit(0)

import openpyxl
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
name = os.path.basename(file_path)
result = {"file": name, "sheets": {}}
sheets = [sheet_filter] if sheet_filter and sheet_filter in wb.sheetnames else wb.sheetnames
for sname in sheets:
    ws = wb[sname]
    rows = []
    for row in ws.iter_rows(values_only=True):
        str_row = [str(c) if c is not None else "" for c in row]
        if any(cell.strip() for cell in str_row):
            rows.append([str(c) if c is not None else "" for c in row])
    if rows:
        result["sheets"][sname] = {"header": rows[0], "rows": rows[1:], "row_count": len(rows)-1, "col_count": len(rows[0])}
wb.close()
json.dump(result, sys.stdout, ensure_ascii=False)
'''


def _find_file(file_name: str) -> Path | None:
    docs = Path(DOCS_PATH)
    if not docs.exists():
        return None
    for f in docs.rglob("*"):
        if f.is_file() and file_name.lower() in f.name.lower():
            if f.suffix.lower() in [".xlsx", ".xls", ".csv"]:
                return f
    return None


def _read_spreadsheet(file_path: Path, sheet_name: str | None = None) -> dict:
    """Read Excel/CSV via subprocess and return as JSON."""
    args = [PYTHON, "-c", _READ_SCRIPT, str(file_path), sheet_name or ""]
    proc = subprocess.run(args, capture_output=True, text=True, timeout=30, stdin=subprocess.DEVNULL)
    if proc.returncode != 0:
        logger.error(f"read error: {proc.stderr[:500]}")
        return {"file": file_path.name, "sheets": {}}
    return json.loads(proc.stdout)


@mcp.tool()
def ping(message: str = "pong") -> str:
    """Test ping.

    Args:
        message: Message to return
    """
    return f"data-analysis alive: {message}"


@mcp.tool()
def analyze_spreadsheet(file_name: str) -> str:
    """Analyzes the structure of an Excel/CSV file (sheet list, column names, row count, data preview).

    Args:
        file_name: File name or keyword (e.g., "sales_report", "sales.xlsx")
    """
    file_path = _find_file(file_name)
    if not file_path:
        return f"Could not find file '{file_name}'. Supported formats: .xlsx, .xls, .csv"

    data = _read_spreadsheet(file_path)

    output = f"**File: {data['file']}**\n\n"
    for sname, sheet in data["sheets"].items():
        output += f"### Sheet: {sname}\n"
        output += f"- Columns ({sheet['col_count']}): {', '.join(sheet['header'])}\n"
        output += f"- Data rows: {sheet['row_count']}\n"
        preview = sheet["rows"][:5]
        if preview:
            output += f"\nPreview:\n"
            output += " | ".join(sheet["header"]) + "\n"
            output += " | ".join(["---"] * len(sheet["header"])) + "\n"
            for row in preview:
                sr = list(row) + [""] * (len(sheet["header"]) - len(row))
                output += " | ".join(sr[:len(sheet["header"])]) + "\n"
        output += "\n"
    return output


@mcp.tool()
def query_data(file_name: str, sheet_name: str = "", column: str = "", operation: str = "list", filter_value: str = "") -> str:
    """Queries and aggregates Excel/CSV data.

    Args:
        file_name: File name or keyword
        sheet_name: Sheet name (leave empty for first sheet)
        column: Target column name (for aggregation/filter)
        operation: Operation to perform (list, sum, avg, max, min, count, filter)
        filter_value: Filter value for filter operation
    """
    file_path = _find_file(file_name)
    if not file_path:
        return f"Could not find file '{file_name}'."

    data = _read_spreadsheet(file_path, sheet_name or None)
    if not data["sheets"]:
        return "No data found."

    sname = sheet_name if sheet_name in data["sheets"] else list(data["sheets"].keys())[0]
    sheet = data["sheets"][sname]
    header, rows = sheet["header"], sheet["rows"]

    col_idx = None
    if column:
        for i, h in enumerate(header):
            if column.lower() in h.lower():
                col_idx = i
                break
        if col_idx is None:
            return f"Could not find column '{column}'. Available: {', '.join(header)}"

    if operation == "filter" and filter_value:
        rows = [r for r in rows if filter_value.lower() in " ".join(str(c) for c in r).lower()]

    if operation in ("list", "filter"):
        output = f"**{data['file']} > {sname}** ({len(rows)} rows)\n\n"
        output += " | ".join(header) + "\n" + " | ".join(["---"] * len(header)) + "\n"
        for row in rows[:20]:
            sr = list(row) + [""] * (len(header) - len(row))
            output += " | ".join(sr[:len(header)]) + "\n"
        if len(rows) > 20:
            output += f"\n... and {len(rows)-20} more rows\n"
        return output

    if col_idx is None:
        return "Aggregation operations require the column parameter."

    values = []
    for row in rows:
        if col_idx < len(row):
            try:
                values.append(float(str(row[col_idx]).replace(",", "").replace("%", "")))
            except (ValueError, TypeError):
                pass
    if not values:
        return f"No numeric data in column '{column}'."

    ops = {
        "sum": lambda v: f"**{column} total:** {sum(v):,.2f} ({len(v)} records)",
        "avg": lambda v: f"**{column} average:** {sum(v)/len(v):,.2f} ({len(v)} records)",
        "max": lambda v: f"**{column} maximum:** {max(v):,.2f}",
        "min": lambda v: f"**{column} minimum:** {min(v):,.2f}",
        "count": lambda v: f"**{column} data count:** {len(v)} records",
    }
    return ops.get(operation, lambda v: f"Unsupported operation: {operation}")(values)


@mcp.tool()
def compare_values(file_name: str, column: str, value1_label: str, value2_label: str, sheet_name: str = "") -> str:
    """Compares and analyzes values from two rows.

    Args:
        file_name: File name
        column: Numeric column name to compare
        value1_label: First row keyword (e.g., "January", "Company A")
        value2_label: Second row keyword (e.g., "June", "Company B")
        sheet_name: Sheet name (leave empty for first)
    """
    file_path = _find_file(file_name)
    if not file_path:
        return f"Could not find file '{file_name}'."

    data = _read_spreadsheet(file_path, sheet_name or None)
    if not data["sheets"]:
        return "No data found."

    sheet = data["sheets"].get(sheet_name, list(data["sheets"].values())[0])
    header, rows = sheet["header"], sheet["rows"]

    col_idx = None
    for i, h in enumerate(header):
        if column.lower() in h.lower():
            col_idx = i
            break
    if col_idx is None:
        return f"Could not find column '{column}'."

    val1 = val2 = None
    for row in rows:
        row_str = " ".join(str(c) for c in row).lower()
        if value1_label.lower() in row_str and val1 is None:
            try: val1 = float(str(row[col_idx]).replace(",","").replace("%",""))
            except: pass
        if value2_label.lower() in row_str and val2 is None:
            try: val2 = float(str(row[col_idx]).replace(",","").replace("%",""))
            except: pass

    if val1 is None: return f"Could not find row '{value1_label}'."
    if val2 is None: return f"Could not find row '{value2_label}'."

    diff = val2 - val1
    pct = ((val2 - val1) / abs(val1)) * 100 if val1 != 0 else 0
    direction = "increase" if diff > 0 else "decrease" if diff < 0 else "no change"

    return (
        f"**{column} comparison: {value1_label} vs {value2_label}**\n\n"
        f"- {value1_label}: {val1:,.2f}\n"
        f"- {value2_label}: {val2:,.2f}\n"
        f"- Difference: {diff:+,.2f} ({direction})\n"
        f"- Change rate: {pct:+.1f}%\n"
    )


if __name__ == "__main__":
    mcp.run(transport="stdio")
