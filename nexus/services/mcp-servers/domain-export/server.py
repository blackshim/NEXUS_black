"""
NEXUS Core MCP: domain-export

domains/{domain_name}/domain_knowledge.json -> Excel(.xlsx) export.
Used by CS teams to review/correct the knowledge DB offline.

Design spec:
- DOMAINS_BASE env var specifies the domains/ folder path
- domain_name parameter specifies the target domain
- Output: domains/{domain_name}/exports/domain_knowledge_YYYYMMDD.xlsx

Tools:
- export_knowledge(domain_name, output_path): JSON -> XLSX conversion
"""

import json
import logging
import os
import subprocess
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    class FastMCP:
        def __init__(self, name): self.name = name
        def tool(self, *a, **kw):
            def dec(fn): return fn
            return dec
        def run(self, **kw): pass

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("nexus.mcp.domain-export")

mcp = FastMCP("nexus-domain-export")
PYTHON = sys.executable
KST = timezone(timedelta(hours=9))

DOMAINS_BASE = os.environ.get("DOMAINS_BASE", "")
if not DOMAINS_BASE:
    logger.warning("DOMAINS_BASE environment variable is not set.")
else:
    logger.info(f"Domains base: {DOMAINS_BASE}")


def _get_knowledge_path(domain_name: str) -> Path:
    return Path(DOMAINS_BASE) / domain_name / "domain_knowledge.json"


_EXPORT_SCRIPT = '''
import json, sys, os

args = json.loads(sys.argv[1])
json_path = args["json_path"]
output_path = args["output_path"]

try:
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
except Exception as e:
    json.dump({"error": f"JSON load failed: {str(e)}"}, sys.stdout, ensure_ascii=False)
    sys.exit(0)

items = data.get("items", [])
meta = data.get("_meta", {})

if not items:
    json.dump({"error": "No items to export."}, sys.stdout, ensure_ascii=False)
    sys.exit(0)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    json.dump({"error": "openpyxl is not installed."}, sys.stdout, ensure_ascii=False)
    sys.exit(0)

all_fields = []
skip_fields = {"usage_stats"}
for item in items:
    for key in item.keys():
        if key not in all_fields and key not in skip_fields:
            all_fields.append(key)
all_fields.extend(["usage_suggested", "usage_resolved", "usage_failed", "usage_success_rate"])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Domain Knowledge"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

for col_idx, field in enumerate(all_fields, 1):
    cell = ws.cell(row=1, column=col_idx, value=field)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row_idx, item in enumerate(items, 2):
    for col_idx, field in enumerate(all_fields, 1):
        if field == "usage_suggested":
            value = item.get("usage_stats", {}).get("suggested", 0)
        elif field == "usage_resolved":
            value = item.get("usage_stats", {}).get("resolved", 0)
        elif field == "usage_failed":
            value = item.get("usage_stats", {}).get("failed", 0)
        elif field == "usage_success_rate":
            rate = item.get("usage_stats", {}).get("success_rate", 0)
            value = f"{rate:.0%}" if isinstance(rate, (int, float)) else str(rate)
        elif field == "tags":
            tags = item.get("tags", [])
            value = ", ".join(tags) if isinstance(tags, list) else str(tags)
        else:
            value = item.get(field, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

for col_idx, field in enumerate(all_fields, 1):
    max_len = len(str(field))
    for row_idx in range(2, min(len(items) + 2, 52)):
        cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
        for line in cell_value.split("\\n"):
            max_len = max(max_len, len(line))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

ws_meta = wb.create_sheet("Metadata")
ws_meta.cell(row=1, column=1, value="Item").font = Font(bold=True)
ws_meta.cell(row=1, column=2, value="Value").font = Font(bold=True)
meta_rows = [
    ("Domain", meta.get("domain", "")),
    ("Schema Version", meta.get("schema_version", "")),
    ("Last Updated", meta.get("last_crystallized", "")),
    ("Total Items", len(items)),
]
stats = meta.get("stats", {})
for key, val in stats.items():
    if key != "total_items":
        meta_rows.append((key, val))
for i, (k, v) in enumerate(meta_rows, 2):
    ws_meta.cell(row=i, column=1, value=k)
    ws_meta.cell(row=i, column=2, value=v)

os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)

result = {"status": "exported", "output_path": output_path,
          "total_items": len(items), "fields": len(all_fields)}
json.dump(result, sys.stdout, ensure_ascii=False)
'''


@mcp.tool()
def export_knowledge(domain_name: str, output_path: str = "") -> str:
    """Exports domain knowledge DB (JSON) to an Excel file.

    Args:
        domain_name: Domain name (e.g., "my-domain")
        output_path: Output file path (default: date-stamped file in domains/{domain}/exports/)
    """
    if not DOMAINS_BASE:
        return "DOMAINS_BASE environment variable is not set."

    knowledge_path = _get_knowledge_path(domain_name)
    if not knowledge_path.exists():
        return f"JSON file not found: {knowledge_path}"

    if not output_path:
        export_dir = str(Path(DOMAINS_BASE) / domain_name / "exports")
        now = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
        output_path = f"{export_dir}/domain_knowledge_{now}.xlsx"

    args = json.dumps({
        "json_path": str(knowledge_path),
        "output_path": output_path
    })

    proc = subprocess.run(
        [PYTHON, "-c", _EXPORT_SCRIPT, args],
        capture_output=True, text=True, timeout=60,
        stdin=subprocess.DEVNULL
    )

    if proc.returncode != 0:
        return f"Export failed: {proc.stderr.strip()[:200]}"

    resp = json.loads(proc.stdout)

    if "error" in resp:
        return f"Export failed: {resp['error']}"

    return (
        f"Excel export complete!\n"
        f"- Domain: {domain_name}\n"
        f"- File: {resp.get('output_path')}\n"
        f"- Items: {resp.get('total_items')}\n"
        f"- Columns: {resp.get('fields')}"
    )


if __name__ == "__main__":
    mcp.run(transport="stdio")
