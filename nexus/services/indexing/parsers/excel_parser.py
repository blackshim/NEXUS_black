"""Excel/CSV parser + irregular detection + LLM structure inference (Tier 0)."""

import json
import logging
import os
from pathlib import Path

import httpx

from .base import BaseParser, ParsedDocument, ParsedPage

logger = logging.getLogger("nexus.parser.excel")


def _load_excel_config() -> dict:
    """Load Excel parser related configuration."""
    try:
        from utils.config_loader import get_indexing_config
        return get_indexing_config().get("excel", {})
    except Exception:
        return {}


def _load_llm_config() -> dict:
    """Load LLM configuration."""
    try:
        from utils.config_loader import get_llm_config
        return get_llm_config()
    except Exception:
        return {}


def _load_file_encodings() -> list[str]:
    """Load file encoding fallback list."""
    try:
        from utils.config_loader import get_indexing_config
        return get_indexing_config().get("file_encodings", ["utf-8", "cp949", "euc-kr", "latin-1"])
    except Exception:
        return ["utf-8", "cp949", "euc-kr", "latin-1"]


# === LLM Prompt ===
STRUCTURE_INFERENCE_PROMPT = """Below is the top {n} rows of data from an Excel sheet. Please analyze the structure of this sheet.

## Data
```
{sample}
```

## Terminology
- **header_row**: The row containing column names. Example: "Error Code | Description | Action"
- **data_start_row**: The row where actual data values begin. Often the row immediately after header_row.
- **title_rows**: Rows containing metadata such as sheet title, subtitle, creation date, etc. Usually located in rows 1-3 and occupy only a small portion of the overall data. **Never include data rows in title_rows.**
- **column_names**: List of column names

## Analysis Request
Please respond in the following JSON format. Output only JSON with no other text.

```json
{{
  "header_row": <row number where header is located (1-based, null if none)>,
  "data_start_row": <row number where actual data starts>,
  "title_rows": [<list of row numbers for title/memo rows. Usually 0-3>],
  "sub_tables": [
    {{"start_row": <start>, "end_row": <end>, "description": "<description>"}}
  ],
  "column_names": [<list of inferred column names>],
  "notes": "<brief description of the structure>"
}}
```"""


def _detect_irregular(ws, rows_with_types: list[list], empty_row_scan_depth: int) -> dict:
    """Detect whether a sheet is irregular. Irregular if any signal is present.

    Signals:
    1. Merged cells exist
    2. Row 1 is empty or top N rows contain title/memo (header position unclear)
    3. Empty row in the middle of data (possible multiple tables)
    4. Mixed number/text types in the same column
    """
    signals = {}

    # --- Signal 1: Merged cells ---
    try:
        merged = list(ws.merged_cells.ranges)
        if merged:
            signals["merged_cells"] = len(merged)
    except Exception:
        pass

    if not rows_with_types:
        return signals

    # --- Signal 2: Row 1 empty or top rows contain title/memo ---
    first_row_values = rows_with_types[0] if rows_with_types else []
    non_empty_first = [c for c in first_row_values if c[0].strip()]
    if len(non_empty_first) <= 1 and len(rows_with_types) > 1:
        # Row 1 has 1 or fewer cells -> likely a title
        signals["title_in_first_row"] = True

    # --- Signal 3: Empty row in the middle of data ---
    scan_depth = min(empty_row_scan_depth, len(rows_with_types))
    data_started = False
    for i in range(scan_depth):
        row_values = [c[0] for c in rows_with_types[i]]
        has_data = any(v.strip() for v in row_values)
        if has_data:
            data_started = True
        elif data_started:
            # Empty row after data started -> possible multiple tables
            signals["empty_row_in_data"] = i + 1  # Row number (1-based)
            break

    # --- Signal 4: Mixed number/text types in the same column ---
    if len(rows_with_types) >= 3:
        num_cols = max(len(row) for row in rows_with_types[:scan_depth])
        for col_idx in range(num_cols):
            col_types = set()
            for row in rows_with_types[1:scan_depth]:  # Exclude row 1 (possible header)
                if col_idx < len(row):
                    _, cell_type = row[col_idx]
                    if cell_type != "empty":
                        col_types.add(cell_type)
            if "number" in col_types and "text" in col_types:
                signals["mixed_types_in_column"] = col_idx + 1  # Column number (1-based)
                break

    return signals


def _call_ollama(ollama_url: str, model: str, prompt: str, llm_params: dict) -> str | None:
    """Call Ollama model once. Returns response text on success, None on failure."""
    timeout = llm_params.get("timeout", 60)
    temperature = llm_params.get("temperature", 0.1)
    max_tokens = llm_params.get("max_tokens", 1024)

    try:
        with httpx.Client(timeout=float(timeout)) as client:
            resp = client.post(
                f"{ollama_url}/api/chat",
                json={
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "stream": False,
                    "think": False,  # Compatible with thinking models (qwen3, etc.)
                    "options": {"temperature": temperature, "num_predict": max_tokens},
                },
            )
            resp.raise_for_status()
            content = resp.json().get("message", {}).get("content", "")
            if content.strip():
                return content
    except Exception as e:
        logger.warning(f"Ollama call failed for {model}: {e}")
    return None


def _call_llm(prompt: str) -> str | None:
    """Request structure inference using LLM settings from nexus.config.yaml.

    Ollama: /api/chat + think=false (compatible with thinking models)
    On configured model failure: auto-fallback to other available Ollama models
    Final fallback: OpenAI-compatible API
    """
    llm_config = _load_llm_config()

    # Try offline (Ollama) first
    offline = llm_config.get("offline", {})
    endpoint = offline.get("endpoint", "http://localhost:11434")
    model = offline.get("model", "qwen3:4b")
    offline_params = {
        "timeout": offline.get("timeout", 60),
        "temperature": offline.get("temperature", 0.1),
        "max_tokens": offline.get("max_tokens", 1024),
    }

    # Access Ollama via host.docker.internal from inside Docker
    ollama_url = os.environ.get("OLLAMA_URL", endpoint)

    # Try configured model
    result = _call_ollama(ollama_url, model, prompt, offline_params)
    if result and _parse_llm_response(result) is not None:
        return result

    # Configured model failed -> try other available models
    if result is None or _parse_llm_response(result) is None:
        logger.warning(f"Model {model} failed for structure inference, trying alternatives")
        try:
            with httpx.Client(timeout=5.0) as client:
                tags_resp = client.get(f"{ollama_url}/api/tags")
                available = [m["name"] for m in tags_resp.json().get("models", [])]
        except Exception:
            available = []

        for alt_model in available:
            if alt_model == model:
                continue
            alt_result = _call_ollama(ollama_url, alt_model, prompt, offline_params)
            if alt_result and _parse_llm_response(alt_result) is not None:
                logger.info(f"Fallback model {alt_model} succeeded")
                return alt_result

    # Online (OpenAI-compatible) fallback
    online = llm_config.get("online", {})
    api_key_env = online.get("api_key_env", "OPENAI_API_KEY")
    api_key = os.environ.get(api_key_env)
    if not api_key:
        logger.warning("No LLM available for structure inference")
        return None

    online_model = online.get("model", "gpt-4o-mini")
    online_timeout = online.get("timeout", 60)
    online_temperature = online.get("temperature", 0.1)
    online_max_tokens = online.get("max_tokens", 1024)

    try:
        with httpx.Client(timeout=float(online_timeout)) as client:
            resp = client.post(
                "https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}"},
                json={
                    "model": online_model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": online_temperature,
                    "max_tokens": online_max_tokens,
                },
            )
            resp.raise_for_status()
            return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"Online LLM call also failed: {e}")
        return None


def _parse_llm_response(response: str) -> dict | None:
    """Extract JSON from LLM response.

    Tries multiple strategies to find JSON blocks:
    1. ```json ... ``` block
    2. ``` ... ``` block
    3. First { ... } object in text
    """
    if not response:
        return None

    text = response.strip()

    # Strategy 1: ```json ... ``` block
    if "```json" in text:
        try:
            start = text.index("```json") + 7
            end = text.index("```", start)
            return json.loads(text[start:end].strip())
        except (ValueError, json.JSONDecodeError):
            pass

    # Strategy 2: ``` ... ``` block
    if "```" in text:
        try:
            start = text.index("```") + 3
            end = text.index("```", start)
            candidate = text[start:end].strip()
            if candidate.startswith("{"):
                return json.loads(candidate)
        except (ValueError, json.JSONDecodeError):
            pass

    # Strategy 3: Direct { ... } object extraction from text
    try:
        brace_start = text.index("{")
        # Find last }
        brace_end = text.rindex("}") + 1
        candidate = text[brace_start:brace_end]
        return json.loads(candidate)
    except (ValueError, json.JSONDecodeError):
        pass

    logger.warning(f"Failed to parse LLM JSON response: {text[:200]}")
    return None


class ExcelParser(BaseParser):
    """openpyxl-based Excel parser + irregular detection + LLM structure inference.

    Regular: Row 1 header + data -> standard markdown conversion
    Irregular: Merged cells, title rows, empty rows detected -> request LLM structure inference
    """

    supported_extensions = [".xlsx", ".xls", ".csv"]

    def __init__(self):
        excel_config = _load_excel_config()
        self._sample_rows = excel_config.get("sample_rows", 20)
        self._empty_row_scan_depth = excel_config.get("empty_row_scan_depth", 50)
        self._title_rows_sanity_threshold = excel_config.get("title_rows_sanity_threshold", 0.5)

    def parse(self, file_path: Path) -> ParsedDocument:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv(file_path)
        if suffix == ".xls":
            return self._parse_xls(file_path)
        return self._parse_excel(file_path)

    def _read_rows_with_types(self, ws) -> list[list[tuple[str, str]]]:
        """Read all rows of a sheet as a list of (string_value, type) tuples."""
        rows = []
        for row in ws.iter_rows():
            typed_row = []
            for cell in row:
                if cell.value is None:
                    typed_row.append(("", "empty"))
                elif isinstance(cell.value, (int, float)):
                    typed_row.append((str(cell.value), "number"))
                else:
                    typed_row.append((str(cell.value), "text"))
            if any(c[0].strip() for c in typed_row):
                rows.append(typed_row)
        return rows

    def _format_sample_for_llm(self, rows_with_types: list[list[tuple[str, str]]]) -> str:
        """Format sheet sample as text for LLM."""
        sample = rows_with_types[:self._sample_rows]
        lines = []
        for i, row in enumerate(sample, 1):
            values = [c[0] for c in row]
            lines.append(f"Row{i}: {' | '.join(values)}")
        return "\n".join(lines)

    def _parse_irregular_sheet(
        self, sheet_name: str, rows_with_types: list[list[tuple[str, str]]], signals: dict
    ) -> ParsedPage:
        """Parse irregular sheet using LLM structure inference."""
        sample_text = self._format_sample_for_llm(rows_with_types)
        prompt = STRUCTURE_INFERENCE_PROMPT.format(
            n=min(self._sample_rows, len(rows_with_types)), sample=sample_text
        )

        llm_response = _call_llm(prompt)
        structure = _parse_llm_response(llm_response)

        if structure:
            return self._apply_structure(sheet_name, rows_with_types, structure, signals)
        else:
            # LLM failure -> fallback to existing logic
            logger.warning(f"LLM structure inference failed for '{sheet_name}', using fallback")
            return self._parse_regular_sheet(sheet_name, rows_with_types, signals)

    def _apply_structure(
        self, sheet_name: str, rows_with_types: list[list[tuple[str, str]]],
        structure: dict, signals: dict
    ) -> ParsedPage:
        """Apply LLM-inferred structure to generate ParsedPage."""
        header_row = structure.get("header_row")
        data_start = structure.get("data_start_row", 1)
        column_names = structure.get("column_names", [])
        title_rows = structure.get("title_rows", [])
        notes = structure.get("notes", "")

        total_rows = len(rows_with_types)

        # --- LLM output correction ---
        # If title_rows is >= threshold of total rows, LLM error -> keep only rows before data_start
        if title_rows and len(title_rows) >= total_rows * self._title_rows_sanity_threshold:
            logger.warning(
                f"LLM title_rows sanity check: {len(title_rows)}/{total_rows} rows marked as title, "
                f"keeping only rows before data_start={data_start}"
            )
            title_rows = [r for r in title_rows if r < data_start]

        # Correct if data_start exceeds total row count
        if data_start > total_rows:
            data_start = 2 if header_row == 1 else 1

        # Collect title row text
        title_text = ""
        for row_num in title_rows:
            if 0 < row_num <= total_rows:
                row_values = [c[0] for c in rows_with_types[row_num - 1]]
                non_empty = [v for v in row_values if v.strip()]
                if non_empty:
                    title_text += " ".join(non_empty) + "\n"

        # Determine header
        if column_names:
            header = column_names
        elif header_row and 0 < header_row <= total_rows:
            header = [c[0] for c in rows_with_types[header_row - 1]]
        else:
            # If header cannot be inferred, use Col1, Col2...
            max_cols = max(len(r) for r in rows_with_types) if rows_with_types else 0
            header = [f"Col{i+1}" for i in range(max_cols)]

        # Extract data rows (after data_start, excluding title/header rows)
        data_rows = []
        for i in range(data_start - 1, total_rows):
            row_num = i + 1
            if row_num in title_rows:
                continue
            if header_row is not None and row_num == header_row:
                continue
            row_values = [c[0] for c in rows_with_types[i]]
            if any(v.strip() for v in row_values):
                data_rows.append(row_values)

        # Generate markdown table
        separator = ["---"] * len(header)
        table_lines = [" | ".join(header), " | ".join(separator)]
        for row in data_rows:
            padded = row + [""] * (len(header) - len(row))
            table_lines.append(" | ".join(padded[:len(header)]))
        table_text = "\n".join(table_lines)

        # Full text
        text_parts = []
        if title_text.strip():
            text_parts.append(title_text.strip())
        for row in data_rows:
            non_empty = [c for c in row if c.strip()]
            if non_empty:
                text_parts.append(" ".join(non_empty))
        full_text = "\n".join(text_parts)

        return ParsedPage(
            page_or_sheet=sheet_name,
            text=full_text,
            tables=[table_text],
            metadata={
                "row_count": len(data_rows),
                "irregular": True,
                "signals": signals,
                "llm_structure": structure,
                "notes": notes,
            }
        )

    def _parse_regular_sheet(
        self, sheet_name: str, rows_with_types: list[list[tuple[str, str]]],
        signals: dict | None = None
    ) -> ParsedPage:
        """Parse regular sheet using existing logic (assuming row 1 header)."""
        rows = [[c[0] for c in row] for row in rows_with_types]

        if len(rows) >= 2:
            header = rows[0]
            separator = ["---"] * len(header)
            table_lines = [" | ".join(header), " | ".join(separator)]
            for row in rows[1:]:
                padded = row + [""] * (len(header) - len(row))
                table_lines.append(" | ".join(padded[:len(header)]))
            table_text = "\n".join(table_lines)
        else:
            table_text = " | ".join(rows[0]) if rows else ""

        text_parts = []
        for row in rows:
            non_empty = [c for c in row if c.strip()]
            if non_empty:
                text_parts.append(" ".join(non_empty))
        full_text = "\n".join(text_parts)

        metadata = {"row_count": len(rows), "irregular": False}
        if signals:
            metadata["signals"] = signals

        return ParsedPage(
            page_or_sheet=sheet_name,
            text=full_text,
            tables=[table_text],
            metadata=metadata
        )

    def _parse_xls(self, file_path: Path) -> ParsedDocument:
        """Parse legacy .xls files using xlrd (regular processing only, limited merged cell detection)."""
        import xlrd

        wb = xlrd.open_workbook(str(file_path))
        pages = []

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            if ws.nrows == 0:
                continue

            rows = []
            for row_idx in range(ws.nrows):
                row_values = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is None or cell.value == "":
                        row_values.append("")
                    else:
                        row_values.append(str(cell.value))
                if any(v.strip() for v in row_values):
                    rows.append(row_values)

            if not rows:
                continue

            # Generate markdown table
            if len(rows) >= 2:
                header = rows[0]
                separator = ["---"] * len(header)
                table_lines = [" | ".join(header), " | ".join(separator)]
                for row in rows[1:]:
                    padded = row + [""] * (len(header) - len(row))
                    table_lines.append(" | ".join(padded[:len(header)]))
                table_text = "\n".join(table_lines)
            else:
                table_text = " | ".join(rows[0]) if rows else ""

            text_parts = []
            for row in rows:
                non_empty = [c for c in row if c.strip()]
                if non_empty:
                    text_parts.append(" ".join(non_empty))
            full_text = "\n".join(text_parts)

            pages.append(ParsedPage(
                page_or_sheet=sheet_name,
                text=full_text,
                tables=[table_text],
                metadata={"row_count": len(rows), "irregular": False}
            ))

        return ParsedDocument(
            file_path=str(file_path),
            file_type="xls",
            pages=pages,
        )

    def _parse_excel(self, file_path: Path) -> ParsedDocument:
        import openpyxl

        # Open in normal mode for irregular detection (need access to merged_cells)
        wb = openpyxl.load_workbook(str(file_path), read_only=False, data_only=True)
        pages = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Read rows with type information
            rows_with_types = self._read_rows_with_types(ws)
            if not rows_with_types:
                continue

            # Detect irregular
            signals = _detect_irregular(ws, rows_with_types, self._empty_row_scan_depth)

            if signals:
                logger.info(f"Irregular sheet detected: '{sheet_name}' — signals: {signals}")
                page = self._parse_irregular_sheet(sheet_name, rows_with_types, signals)
            else:
                page = self._parse_regular_sheet(sheet_name, rows_with_types)

            pages.append(page)

        wb.close()

        return ParsedDocument(
            file_path=str(file_path),
            file_type="xlsx",
            pages=pages,
        )

    def _parse_csv(self, file_path: Path) -> ParsedDocument:
        import csv

        encodings = _load_file_encodings()

        rows = []
        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = [row for row in reader]
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if not rows:
            return ParsedDocument(file_path=str(file_path), file_type="csv")

        # Markdown table
        header = rows[0]
        separator = ["---"] * len(header)
        table_lines = [" | ".join(header), " | ".join(separator)]
        for row in rows[1:]:
            padded = row + [""] * (len(header) - len(row))
            table_lines.append(" | ".join(padded[:len(header)]))
        table_text = "\n".join(table_lines)

        full_text = "\n".join(" ".join(r) for r in rows if any(c.strip() for c in r))

        pages = [ParsedPage(
            page_or_sheet="Sheet1",
            text=full_text,
            tables=[table_text],
            metadata={"row_count": len(rows)}
        )]

        return ParsedDocument(
            file_path=str(file_path),
            file_type="csv",
            pages=pages,
        )
